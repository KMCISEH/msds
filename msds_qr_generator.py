import os
import urllib.parse
import qrcode
import PIL
from PIL import Image as PILImage
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage

# 1. 경로 설정
# 대상 디렉토리 (MSDS PDF 파일들이 위치한 경로)
TARGET_DIR = r"Z:\안전환경팀\박봉육\공정별 유해위험물질 MSDS"
# QR 코드 이미지 파일들을 저장할 하위 폴더
QR_DIR = os.path.join(TARGET_DIR, "QR_Codes")
# 생성할 엑셀 파일 경로
EXCEL_PATH = os.path.join(TARGET_DIR, "공정별_유해위험물질_MSDS_QR목록.xlsx")

# 2. QR 코드 링크의 기본 주소 (GitHub 리포지토리 raw 파일 경로 설정)
# 모바일 기기 브라우저가 리다이렉트 없이 바로 PDF 파일 스트림을 받아 기본 PDF 뷰어로 연결할 수 있도록 직통 주소(raw.githubusercontent.com)로 지정합니다.
BASE_URL = "https://raw.githubusercontent.com/KMCISEH/msds/main/"

def clean_filename_for_url(filename):
    # 파일명을 URL 인코딩하여 링크 생성 시 오류가 없도록 처리
    return urllib.parse.quote(filename)

def main():
    print("MSDS QR 코드 및 엑셀 생성 작업을 시작합니다...")
    
    # 대상 디렉토리 확인
    if not os.path.exists(TARGET_DIR):
        print(f"오류: 대상 폴더가 존재하지 않습니다: {TARGET_DIR}")
        return
        
    # QR_Codes 폴더 생성
    if not os.path.exists(QR_DIR):
        os.makedirs(QR_DIR)
        print(f"QR 코드 저장 폴더를 생성했습니다: {QR_DIR}")
        
    # PDF 파일 목록 가져오기
    files = os.listdir(TARGET_DIR)
    pdf_files = sorted([f for f in files if f.lower().endswith('.pdf')])
    
    if not pdf_files:
        print("대상 폴더에 PDF 파일이 없습니다.")
        return
        
    print(f"총 {len(pdf_files)}개의 PDF 파일을 찾았습니다.")
    
    # 엑셀 워크북 초기화
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MSDS QR코드 목록"
    
    # 그리드라인 보이기 설정
    ws.views.sheetView[0].showGridLines = True
    
    # 스타일 정의
    font_title = Font(name="맑은 고딕", size=16, bold=True, color="1F4E78")
    font_header = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="맑은 고딕", size=10, color="000000")
    font_link = Font(name="맑은 고딕", size=10, color="0563C1", underline="single")
    
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Title 작성
    ws.merge_cells("A1:E1")
    ws["A1"] = "공정별 유해위험물질 MSDS QR코드 목록"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Table Header 작성
    headers = ["번호", "물질명 (파일명)", "QR 코드", "바로가기 링크", "비고"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    
    ws.row_dimensions[3].height = 30
    
    # 데이터 행 설정
    # 각 열의 너비 설정
    ws.column_dimensions['A'].width = 8   # 번호
    ws.column_dimensions['B'].width = 35  # 물질명
    ws.column_dimensions['C'].width = 18  # QR 코드 (이미지가 드롭될 열)
    ws.column_dimensions['D'].width = 40  # 바로가기 링크
    ws.column_dimensions['E'].width = 15  # 비고
    
    start_row = 4
    for idx, filename in enumerate(pdf_files, 1):
        current_row = start_row + idx - 1
        
        # 파일명에서 확장자 제외한 이름 (물질명)
        substance_name = os.path.splitext(filename)[0]
        
        # URL 생성
        encoded_filename = clean_filename_for_url(filename)
        raw_url = BASE_URL + encoded_filename
        
        # 안드로이드 등 모바일 브라우저에서 다운로드 창 대신 바로 PDF를 볼 수 있도록 Google Docs Viewer 사용
        file_url = f"https://docs.google.com/viewer?url={urllib.parse.quote(raw_url, safe='')}"
        qr_url = file_url
        
        # QR 코드 생성
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=2
        )
        qr.add_data(qr_url)
        qr.make(fit=True)
        
        qr_img = qr.make_image(fill_color="black", back_color="white")
        # 100x100 픽셀로 크기 조정하여 일관성 유지
        qr_img_resized = qr_img.resize((100, 100), PILImage.Resampling.LANCZOS)
        
        # QR 코드 이미지 파일 저장
        qr_filename = f"{substance_name}_QR.png"
        qr_path = os.path.join(QR_DIR, qr_filename)
        qr_img_resized.save(qr_path)
        
        # 엑셀 데이터 입력
        # 1. 번호
        cell_num = ws.cell(row=current_row, column=1, value=idx)
        cell_num.font = font_data
        cell_num.alignment = align_center
        cell_num.border = thin_border
        
        # 2. 물질명
        cell_name = ws.cell(row=current_row, column=2, value=substance_name)
        cell_name.font = Font(name="맑은 고딕", size=10, bold=True)
        cell_name.alignment = align_left
        cell_name.border = thin_border
        
        # 3. QR 코드 이미지 삽입
        ws.row_dimensions[current_row].height = 85  # 이미지 높이(100px = 75pt)보다 넉넉하게 설정
        excel_img = OpenpyxlImage(qr_path)
        
        # 셀 안에서 약간의 여백을 주기 위해 위치 미세 조정 가능 (openpyxl은 기본적으로 셀 좌상단에 이미지 추가)
        ws.add_image(excel_img, f"C{current_row}")
        
        # QR 코드 셀 테두리 및 정렬 스타일 적용을 위한 빈 셀 설정
        cell_qr = ws.cell(row=current_row, column=3)
        cell_qr.border = thin_border
        
        # 4. 바로가기 링크
        cell_link = ws.cell(row=current_row, column=4, value="클릭하여 파일 열기")
        cell_link.hyperlink = file_url
        cell_link.font = font_link
        cell_link.alignment = align_center
        cell_link.border = thin_border
        
        # 5. 비고
        cell_note = ws.cell(row=current_row, column=5, value="")
        cell_note.font = font_data
        cell_note.alignment = align_center
        cell_note.border = thin_border
        
        # Zebra 스트라이핑 (홀수 행 배경색 입히기)
        if idx % 2 == 1:
            cell_num.fill = fill_zebra
            cell_name.fill = fill_zebra
            cell_qr.fill = fill_zebra
            cell_link.fill = fill_zebra
            cell_note.fill = fill_zebra
            
        print(f"[{idx}/{len(pdf_files)}] {substance_name} QR 코드 생성 및 엑셀 삽입 완료")

    # 엑셀 파일 저장
    try:
        wb.save(EXCEL_PATH)
        print(f"\n성공: 엑셀 파일이 저장되었습니다 -> {EXCEL_PATH}")
    except PermissionError:
        print(f"\n오류: 엑셀 파일('{os.path.basename(EXCEL_PATH)}')이 이미 열려 있습니다.")
        print("엑셀 프로그램을 닫고 스크립트를 다시 실행해 주세요.")
        import sys
        sys.exit(1)
    print(f"QR 코드 이미지 폴더 -> {QR_DIR}")

if __name__ == "__main__":
    main()
